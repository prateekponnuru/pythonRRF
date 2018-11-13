from django.shortcuts import render
from django.contrib.auth.forms import UserCreationForm
from django.urls import reverse_lazy
from django.views import generic
from django.core.files.storage import FileSystemStorage
import openpyxl as xl
from graphos.sources.simple import SimpleDataSource as ds
from graphos.renderers.gchart import BarChart

# Create your views here.

class SignUp(generic.CreateView):
    form_class = UserCreationForm
    success_url = reverse_lazy('FileUpload')
    template_name = 'signup.html'


def file_upload(request):
    if request.method == 'POST' and request.FILES['myfile']:
        myfile = request.FILES['myfile']
        fs = FileSystemStorage()
        filename = fs.save(myfile.name, myfile)
        uploaded_file_url = fs.url(filename)
        wb = xl.load_workbook(myfile)
        ws = wb['Sheet1']
        x_ = ws['A2':'A13']
        x = [a[0].value for a in x_]
        y_ = ws['B2':'B13']
        y = [a[0].value for a in y_]
        file_data = [['Month', 'Sales']]
        file_data.extend(list(list(a) for a in list(zip(x, y))))
        graph_data = ds(data=file_data)
        chart = BarChart(graph_data)
        context = {'chart': chart, 'fileuploaded': True}
        return render(request, 'chart.html', context)
    return render(request, 'fileupload.html')



    
                                            